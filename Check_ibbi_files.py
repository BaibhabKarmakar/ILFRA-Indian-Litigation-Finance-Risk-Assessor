"""
check_ibbi_files.py
-------------------
Run this whenever you add new IBBI quarterly files to data/raw/ibbi/
before running the full pipeline.

It scans every file and tells you exactly which sheet contains
resolution data and which contains liquidation data, so you can
verify the parser will find them correctly.

Usage:
    python check_ibbi_files.py
"""

from openpyxl import load_workbook
from pathlib import Path

IBBI_DIR = Path("data/raw/ibbi")

RESOLUTION_KEYWORDS = ["cirps yielding resolution", "cirps yielding"]
LIQUIDATION_KEYWORDS = ["details of closed liquidation", "closed liquidations"]


def find_sheet(wb, keywords):
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 5:
                break
            row_text = " ".join(str(c).lower() for c in row if c is not None)
            if any(kw.lower() in row_text for kw in keywords):
                return sheet_name
    return None


def count_data_rows(wb, sheet_name):
    if not sheet_name:
        return 0
    ws = wb[sheet_name]
    count = 0
    for row in ws.iter_rows(values_only=True):
        val = row[1] if len(row) > 1 else None
        if isinstance(val, (int, float)) and val == int(val):
            count += 1
    return count


def check_dates(wb, sheet_name):
    """Sample first data row to see if dates are datetime or string."""
    if not sheet_name:
        return "N/A"
    ws = wb[sheet_name]
    for row in ws.iter_rows(values_only=True):
        val = row[1] if len(row) > 1 else None
        if isinstance(val, (int, float)):
            date_val = row[4] if len(row) > 4 else None
            if date_val is None:
                return "missing"
            import datetime
            if isinstance(date_val, datetime.datetime):
                return "datetime ✅"
            elif isinstance(date_val, str):
                return f"string ({date_val}) ⚠️"
            else:
                return f"unknown type: {type(date_val)}"
    return "no data rows"


print("=" * 65)
print("IBBI File Checker")
print("=" * 65)

files = sorted(IBBI_DIR.glob("*.xlsx"))
if not files:
    print(f"No .xlsx files found in {IBBI_DIR}")
    exit()

issues = []

for path in files:
    wb = load_workbook(path, read_only=True, data_only=True)
    res_sheet  = find_sheet(wb, RESOLUTION_KEYWORDS)
    liq_sheet  = find_sheet(wb, LIQUIDATION_KEYWORDS)
    res_rows   = count_data_rows(wb, res_sheet)
    liq_rows   = count_data_rows(wb, liq_sheet)
    date_check = check_dates(wb, res_sheet)

    res_status  = f"{res_sheet} ({res_rows} rows)"  if res_sheet  else "NOT FOUND ❌"
    liq_status  = f"{liq_sheet} ({liq_rows} rows)"  if liq_sheet  else "NOT FOUND ❌"

    print(f"\n── {path.name} ──")
    print(f"  Resolution : {res_status}")
    print(f"  Liquidation: {liq_status}")
    print(f"  Date format: {date_check}")

    if not res_sheet:
        issues.append(f"{path.name}: resolution table not found")
    if not liq_sheet:
        issues.append(f"{path.name}: liquidation table not found")
    if res_rows == 0 and res_sheet:
        issues.append(f"{path.name}: resolution sheet found but 0 data rows")
    if liq_rows == 0 and liq_sheet:
        issues.append(f"{path.name}: liquidation sheet found but 0 data rows")

print("\n" + "=" * 65)
if issues:
    print(f"⚠️  {len(issues)} issue(s) found — fix before running pipeline:")
    for issue in issues:
        print(f"   • {issue}")
else:
    print(f"✅  All {len(files)} files look good — safe to run pipeline.")
print("=" * 65)