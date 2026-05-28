# Add to debug_xlsx.py temporarily
from openpyxl import load_workbook
wb = load_workbook("data/raw/ibbi/Q1_2024.xlsx", read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)
for sheet_name in ["Table 9", "Table 10", "Table 11"]:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 4: break
            print(f"  {sheet_name} row[{i}]: {row[:6]}")