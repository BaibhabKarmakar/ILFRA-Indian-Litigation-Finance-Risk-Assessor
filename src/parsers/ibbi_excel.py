"""
src/ingestion/parsers/ibbi_excel.py
------------------------------------
Parses IBBI quarterly newsletter Excel files.
Extracts Table 8 (Resolution Plans) and Table 14 (Closed Liquidations).

This is a format-specific parser — it knows about the internal
structure of IBBI Excel files and nothing else. It returns a
raw DataFrame. All further cleaning happens in ibbi_channel.py.

If IBBI changes their Excel structure in a future quarter,
this is the only file you need to touch.
"""

import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from src.genai.genai_utils import detect_columns


# ── Table detection helpers ───────────────────────────────────────────────────

def _is_data_row(row: tuple) -> bool:
    """
    A real data row has a serial number (integer or float) in
    column index 0 or 1. Handles format variations across quarters.
    """
    for col_idx in [0, 1, 2]:
        val = row[col_idx] if len(row) > col_idx else None
        if val is None:
            continue
        # Accept int, float that is whole number, or string digit
        if isinstance(val, int):
            return True
        if isinstance(val, float) and val == int(val):
            return True
        if isinstance(val, str) and val.strip().isdigit():
            return True
    return False


def _find_table_sheet(wb, keywords: list[str], content_keywords: list[str] = None) -> str | None:
    """
    Finds a sheet whose name contains any of the given keywords AND
    optionally whose content contains content_keywords in the title row.
    Checks keywords in order — first match with valid content wins.
    """
    for kw in keywords:
        for name in wb.sheetnames:
            normalised_name = name.strip().lower().replace(" ", "")
            if kw.lower().replace(" ", "") in normalised_name:
                # If content validation required, check the sheet title row
                if content_keywords:
                    ws = wb[name]
                    sheet_text = ""
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i > 5:
                            break
                        for cell in row:
                            if isinstance(cell, str):
                                sheet_text += cell.lower()
                    # Sheet must contain at least one content keyword
                    if not any(ck.lower() in sheet_text for ck in content_keywords):
                        continue   # wrong sheet — keep searching
                return name
    return None

def _get_col_index(ws, table_name: str) -> dict[str, int]:
    """
    Reads the header row from a sheet and uses GenAI to map headers
    to canonical column names. Returns {canonical_name: col_index}.
    Falls back to empty dict if no headers found.
    """
    headers = []
    header_row_idx = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Header row has strings in most cells and no integer serial in col 1
        non_empty = [c for c in row if c is not None]
        if len(non_empty) >= 3 and not isinstance(row[1], int):
            headers = [str(c).strip() if c is not None else "" for c in row]
            header_row_idx = i
            break

    if not headers:
        print(f"  [ibbi_excel] Could not find header row in {table_name}")
        return {}

    mapping = detect_columns(headers)  # {raw_header → canonical_name}

    # Invert to {canonical_name → col_index}
    col_index = {}
    for raw_header, canonical in mapping.items():
        if raw_header in headers:
            col_index[canonical] = headers.index(raw_header)

    print(f"  [ibbi_excel] {table_name}: mapped {len(col_index)} columns via GenAI")
    return col_index


def _find_sheet_by_content(wb, content_keywords: list[str]) -> str | None:
    """
    Scans ALL sheets and returns the first one whose title row
    contains any of the content_keywords. Completely ignores sheet names.
    Use this when IBBI keeps renumbering tables across quarters.
    """
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 5:
                break
            row_text = " ".join(str(c).lower() for c in row if c is not None)
            if any(kw.lower() in row_text for kw in content_keywords):
                return sheet_name
    return None

# ── Table 8: Resolution Plans ─────────────────────────────────────────────────

def _parse_table8(wb, quarter_label: str) -> pd.DataFrame:
    # "cirps yielding" is specific enough to only match the main resolution table
    # Avoid "resolution plans" alone — it matches FiSP tables too
    sheet_name = _find_sheet_by_content(wb, [
        "cirps yielding resolution",
        "cirps yielding",
    ])
    if not sheet_name:
        # Fallback to name-based search
        sheet_name = _find_table_sheet(
            wb,
            ["Table 8", "Table8", "Table 9", "Table9",
             "Table 5", "Table5", "Resolution Plan"],
            content_keywords=["resolution plan", "cirps yielding"]
        )
    if not sheet_name:
        print(f"  [ibbi_excel] Resolution table not found in {quarter_label} — skipping.")
        return pd.DataFrame()
    
    ws = wb[sheet_name]
    rows = []

    for row in ws.iter_rows(values_only=True):
        if not _is_data_row(row):
            continue

        company    = row[2]
        start_date = row[4]
        res_date   = row[5]
        admitted   = row[7]  if len(row) > 7  else None
        liq_val    = row[8]  if len(row) > 8  else None
        fair_val   = row[9]  if len(row) > 9  else None
        realisable = row[10] if len(row) > 10 else None
        real_pct   = row[11] if len(row) > 11 else None

        if not isinstance(admitted, (int, float)):
            continue
        if not isinstance(realisable, (int, float)):
            continue

        admitted   = float(admitted)
        realisable = float(realisable)

        if isinstance(real_pct, (int, float)):
            realisation_pct = float(real_pct)
        elif admitted > 0:
            realisation_pct = round(realisable / admitted * 100, 2)
        else:
            realisation_pct = np.nan

        rows.append({
            "company_name":      str(company).strip() if company else None,
            "cirp_start_date":   start_date,
            "resolution_date":   res_date,
            "cirp_initiated_by": str(row[6]).strip() if len(row) > 6 and row[6] else None,
            "admitted_claim_cr": admitted,
            "liquidation_value": float(liq_val) if isinstance(liq_val, (int, float)) else np.nan,
            "fair_value":        float(fair_val) if isinstance(fair_val, (int, float)) else np.nan,
            "realisable_amount": realisable,
            "realisation_pct":   realisation_pct,
            "resolution_status": "Resolution Plan Approved",
            "source_table":      sheet_name,
            "quarter":           quarter_label,
        })

    return pd.DataFrame(rows)


def _parse_table14(wb, quarter_label: str) -> pd.DataFrame:
    sheet_name = _find_sheet_by_content(wb, [
        "details of closed liquidation",
        "closed liquidations",
    ])
    if not sheet_name:
        sheet_name = _find_table_sheet(
            wb,
            ["Table 14", "Table14", "Table 15", "Table15",
             "Table 11", "Table11", "Table 10", "Table10",
             "Table 9",  "Table9",  "Closed Liquidation"],
            content_keywords=["closed liquidation", "dissolution"]
        )
    if not sheet_name:
        print(f"  [ibbi_excel] Liquidation table not found in {quarter_label} — skipping.")
        return pd.DataFrame()

    ws = wb[sheet_name]
    rows = []

    for row in ws.iter_rows(values_only=True):
        if not _is_data_row(row):
            continue

        company     = row[2]
        liq_date    = row[3]
        admitted    = row[4]
        liq_val     = row[5]
        distributed = row[7] if len(row) > 7 else None
        dissolution = row[8] if len(row) > 8 else None

        if not isinstance(admitted, (int, float)):
            continue

        admitted = float(admitted)

        if isinstance(distributed, (int, float)):
            distributed     = float(distributed)
            realisation_pct = round(distributed / admitted * 100, 2) if admitted > 0 else np.nan
        else:
            distributed     = np.nan
            realisation_pct = np.nan

        rows.append({
            "company_name":      str(company).strip() if company else None,
            "cirp_start_date":   liq_date,
            "resolution_date":   dissolution,
            "cirp_initiated_by": None,
            "admitted_claim_cr": admitted,
            "liquidation_value": float(liq_val) if isinstance(liq_val, (int, float)) else np.nan,
            "fair_value":        np.nan,
            "realisable_amount": distributed,
            "realisation_pct":   realisation_pct,
            "resolution_status": "Liquidation Order",
            "source_table":      sheet_name,
            "quarter":           quarter_label,
        })

    return pd.DataFrame(rows)

# ── Public API ────────────────────────────────────────────────────────────────

def parse(path: Path) -> pd.DataFrame:
    """
    Main entry point for the Excel parser.
    Called by ibbi_channel.py when it detects a .xlsx file.

    Opens the workbook, extracts Table 8 and Table 14,
    and returns them stacked into a single DataFrame.

    Parameters
    ----------
    path : Path
        Full path to a single IBBI quarterly .xlsx file.

    Returns
    -------
    pd.DataFrame with columns:
        company_name, cirp_start_date, resolution_date,
        cirp_initiated_by, admitted_claim_cr, liquidation_value,
        fair_value, realisable_amount, realisation_pct,
        resolution_status, source_table, quarter
    """
    quarter_label = path.stem

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [ibbi_excel] Could not open {path.name}: {e}")
        return pd.DataFrame()

    t8  = _parse_table8(wb, quarter_label)
    t14 = _parse_table14(wb, quarter_label)

    print(f"  [ibbi_excel] {path.name} → "
          f"{len(t8)} resolution rows, {len(t14)} liquidation rows")

    frames = [f for f in [t8, t14] if not f.empty]
    if not frames:
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True)